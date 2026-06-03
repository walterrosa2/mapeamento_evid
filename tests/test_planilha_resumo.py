from openpyxl import load_workbook

from src import planilha


RESUMO = """TIPO DE AÇÃO: Reclamação Trabalhista
PARTES:
  - Polo Ativo: João da Silva
  - Polo Passivo: Empresa XYZ Ltda
OBJETO: Cobrança de horas extras
VALORES EM DISPUTA: R$ 50.000,00"""


def _setup_runs(tmp_path, monkeypatch):
    runs = tmp_path / "runs"
    monkeypatch.setattr(planilha, "CAMINHO_RUNS", str(runs))


def test_resumo_e_primeira_aba_e_evidencias_intactas(tmp_path, monkeypatch):
    _setup_runs(tmp_path, monkeypatch)
    run_id = "abc123"

    planilha.inicializar_planilha(run_id)
    evidencias = [
        {"Tipo de Evidência": "Contrato", "Trecho": "t1", "Conteúdo": "c1", "Resumo": "r1", "Referência": "Pág. 1"},
        {"Tipo de Evidência": "Nota Fiscal", "Trecho": "t2", "Conteúdo": "c2", "Resumo": "r2", "Referência": "Pág. 2"},
    ]
    planilha.adicionar_linhas_excel(evidencias, run_id=run_id, arquivo_origem="proc.txt")

    planilha.escrever_resumo_primeira_aba(run_id, RESUMO)

    caminho = planilha.get_caminho_excel(run_id)
    wb = load_workbook(caminho)

    # A aba Resumo deve ser a PRIMEIRA
    assert wb.sheetnames[0] == planilha.NOME_ABA_RESUMO

    # Cabeçalho Campo | Valor
    ws = wb[planilha.NOME_ABA_RESUMO]
    assert ws["A1"].value == "Campo"
    assert ws["B1"].value == "Valor"

    # As evidências continuam acessíveis e completas
    df = planilha.ler_evidencias_df(caminho)
    assert len(df) == 2
    assert set(df["Tipo de Evidência"]) == {"Contrato", "Nota Fiscal"}


def test_resumo_vazio_nao_cria_aba(tmp_path, monkeypatch):
    _setup_runs(tmp_path, monkeypatch)
    run_id = "vazio1"
    planilha.inicializar_planilha(run_id)
    planilha.escrever_resumo_primeira_aba(run_id, "")

    wb = load_workbook(planilha.get_caminho_excel(run_id))
    assert planilha.NOME_ABA_RESUMO not in wb.sheetnames


def test_parsear_resumo_campos_e_linhas_livres():
    pares = planilha._parsear_resumo(RESUMO)
    assert ("TIPO DE AÇÃO", "Reclamação Trabalhista") in pares
    assert ("- Polo Ativo", "João da Silva") in pares
    # Linha 'PARTES:' tem valor vazio mas campo preenchido
    assert ("PARTES", "") in pares
