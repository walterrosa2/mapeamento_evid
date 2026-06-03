import os
from datetime import datetime, timezone
import pandas as pd
from config import CAMINHO_SAIDA, CAMINHO_RUNS

# ---------------------------------------------------------------------------
# Schema
# ---------------------------------------------------------------------------
COLUNAS_EVIDENCIA = [
    "Tipo de Evidência",
    "Trecho",
    "Conteúdo",
    "Resumo",
    "Referência",
]

COLUNAS_META = [
    "Run ID",
    "Arquivo Origem",
    "Data Processamento",
]

COLUNAS_PADRAO = COLUNAS_EVIDENCIA + COLUNAS_META

NOME_ARQUIVO_EXCEL = "evidencias.xlsx"


# ---------------------------------------------------------------------------
# Caminhos
# ---------------------------------------------------------------------------

def get_caminho_excel(run_id: str) -> str:
    pasta = os.path.join(CAMINHO_RUNS, run_id)
    os.makedirs(pasta, exist_ok=True)
    return os.path.join(pasta, NOME_ARQUIVO_EXCEL)


# ---------------------------------------------------------------------------
# Criação / inicialização
# ---------------------------------------------------------------------------

def inicializar_planilha(run_id: str) -> str:
    """Cria Excel isolado para a run. Retorna caminho do arquivo."""
    caminho = get_caminho_excel(run_id)
    if not os.path.exists(caminho):
        df = pd.DataFrame(columns=COLUNAS_PADRAO)
        df.to_excel(caminho, index=False)
    return caminho


# ---------------------------------------------------------------------------
# Inserção
# ---------------------------------------------------------------------------

def _preparar_linha(dados_linha: dict, run_id: str, arquivo_origem: str, ts: str) -> dict:
    row = {k: v for k, v in dados_linha.items() if k in COLUNAS_EVIDENCIA}
    for col in COLUNAS_EVIDENCIA:
        row.setdefault(col, "")
    row["Run ID"] = run_id
    row["Arquivo Origem"] = arquivo_origem
    row["Data Processamento"] = ts
    return row


def adicionar_linhas_excel(lista_dados: list[dict], run_id: str, arquivo_origem: str = "") -> int:
    """
    Escreve todas as evidências de um bloco de uma vez (batch).
    Muito mais eficiente que chamar linha a linha.
    Retorna quantidade de linhas adicionadas após deduplicação.
    """
    if not lista_dados:
        return 0

    caminho = get_caminho_excel(run_id)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")

    novas_linhas = [_preparar_linha(d, run_id, arquivo_origem, ts) for d in lista_dados]
    df_novas = pd.DataFrame(novas_linhas)[COLUNAS_PADRAO]

    df_existente = pd.read_excel(caminho) if os.path.exists(caminho) else pd.DataFrame(columns=COLUNAS_PADRAO)
    antes = len(df_existente)

    df_atualizado = pd.concat([df_existente, df_novas], ignore_index=True)
    df_atualizado.dropna(how="all", inplace=True)
    df_atualizado.drop_duplicates(subset=COLUNAS_EVIDENCIA, keep="first", inplace=True)
    df_atualizado.to_excel(caminho, index=False)

    return len(df_atualizado) - antes


def adicionar_linha_excel(dados_linha: dict, run_id: str, arquivo_origem: str = "") -> None:
    """Compatibilidade: chama a versão batch com uma única linha."""
    adicionar_linhas_excel([dados_linha], run_id, arquivo_origem)


def ler_evidencias_df(caminho: str) -> pd.DataFrame:
    """Lê a aba de evidências de forma resiliente.

    Como a aba 'Resumo' pode estar na primeira posição, seleciona a primeira aba
    que NÃO seja o resumo. Compatível com Excels antigos (aba única 'Sheet1').
    """
    xls = pd.ExcelFile(caminho)
    abas = [s for s in xls.sheet_names if s != NOME_ABA_RESUMO]
    aba = abas[0] if abas else xls.sheet_names[0]
    return pd.read_excel(xls, sheet_name=aba)


# ---------------------------------------------------------------------------
# Retrocompatibilidade (legado — main.py CLI)
# ---------------------------------------------------------------------------

def inicializar_planilha_legado():
    """Mantém compatibilidade com main.py que não usa run_id."""
    caminho_arquivo = os.path.join(CAMINHO_SAIDA, "evidencias_extraidas.xlsx")
    if not os.path.exists(caminho_arquivo):
        df = pd.DataFrame(columns=COLUNAS_EVIDENCIA)
        df.to_excel(caminho_arquivo, index=False)
    return caminho_arquivo


def adicionar_linha_excel_legado(dados_linha: dict):
    """Wrapper legado para main.py CLI (arquivo único acumulativo)."""
    caminho = os.path.join(CAMINHO_SAIDA, "evidencias_extraidas.xlsx")
    dados_filtrados = {k: v for k, v in dados_linha.items() if k in COLUNAS_EVIDENCIA}
    for col in COLUNAS_EVIDENCIA:
        dados_filtrados.setdefault(col, "")
    nova_linha = pd.DataFrame([dados_filtrados])[COLUNAS_EVIDENCIA]
    df_existente = pd.read_excel(caminho) if os.path.exists(caminho) else pd.DataFrame(columns=COLUNAS_EVIDENCIA)
    df_atualizado = pd.concat([df_existente, nova_linha], ignore_index=True)
    df_atualizado.dropna(how="all", inplace=True)
    df_atualizado.drop_duplicates(subset=COLUNAS_EVIDENCIA, keep="first", inplace=True)
    df_atualizado.to_excel(caminho, index=False)


NOME_ABA_RESUMO = "Resumo"


def _parsear_resumo(resumo: str) -> list[tuple[str, str]]:
    """Converte o texto do resumo em pares (Campo, Valor) para a planilha.

    Linhas no formato 'CHAVE: valor' viram duas colunas; demais linhas
    não-vazias vão inteiras para a coluna de valor (Campo vazio).
    """
    linhas: list[tuple[str, str]] = []
    for linha in resumo.splitlines():
        texto = linha.rstrip()
        if not texto.strip():
            continue
        if ":" in texto:
            campo, valor = texto.split(":", 1)
            linhas.append((campo.strip(), valor.strip()))
        else:
            linhas.append(("", texto.strip()))
    return linhas


def escrever_resumo_primeira_aba(run_id: str, resumo: str) -> None:
    """Grava o resumo do processo (SAC) como PRIMEIRA aba ('Resumo') do Excel da run.

    Deve ser chamada APÓS todas as escritas de evidências, pois
    `adicionar_linhas_excel` reescreve o arquivo inteiro e apagaria abas extras.
    Best-effort: qualquer falha apenas gera warning e não derruba o pipeline.
    """
    if not resumo or not resumo.strip():
        return

    caminho = get_caminho_excel(run_id)
    if not os.path.exists(caminho):
        return

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font

        wb = load_workbook(caminho)
        if NOME_ABA_RESUMO in wb.sheetnames:
            del wb[NOME_ABA_RESUMO]

        ws = wb.create_sheet(NOME_ABA_RESUMO, 0)
        ws.append(["Campo", "Valor"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)

        for campo, valor in _parsear_resumo(resumo):
            ws.append([campo, valor])

        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 90
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        wb.save(caminho)
    except Exception as exc:
        from loguru import logger
        logger.warning(f"Falha ao escrever aba '{NOME_ABA_RESUMO}' (best-effort): {exc}")
